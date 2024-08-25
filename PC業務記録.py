import threading
import TkEasyGUI as eg
import time
import win32process, win32api, win32gui, win32con
from pathlib import Path
import psutil
import getpass
import datetime
import openpyxl
import pywintypes

TITLE = "PC業務記録"
desktop_dir = Path.home() / "Desktop"
pitch_time = 10  # 記録間隔


def get_active_window_info():
    dt_now = datetime.datetime.now()  # 現在日付時刻
    # フォアグラウンドウィンドウのハンドルを取得
    hwnd = win32gui.GetForegroundWindow()
    # ウィンドウタイトルを取得
    title = win32gui.GetWindowText(hwnd)
    # ウィンドウタイトル取得
    print("Active Window:", title)
    _, pid = win32process.GetWindowThreadProcessId(hwnd)
    if hwnd != 0 or pid != 0:
        try:
            process = psutil.Process(pid)
            exe_NAME = process.name()
        except pywintypes.error:
            print("error")
            exe_NAME = ""
    # アイドル時間
    idl_time = (win32api.GetTickCount() - win32api.GetLastInputInfo()) / 1000.0
    if idl_time < pitch_time:
        # 短時間の場合 0 とする
        idl_time = 0
    return hwnd, dt_now, title, exe_NAME, idl_time


def record_loop():
    # 初期設定
    last_hwnd = 0
    pc_NAME = win32api.GetComputerName()  # コンピュータ名
    login_id = getpass.getuser()  # ログオンユーザー名
    yyyymmdd = datetime.datetime.now().strftime("%Y%m%d")  # 日付
    if not (desktop_dir / "PC業務記録データ").exists():
        (desktop_dir / "PC業務記録データ").mkdir()
    out_file = f"{desktop_dir}\PC業務記録データ\{pc_NAME}_{yyyymmdd}.xlsx"  # ファイル名
    sheet_name = TITLE  # シート名
    # 出力ファイルが無ければ新規作成
    if (Path(out_file)).exists() == False:

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.cell(row=1, column=1).value = f"開始時間({pitch_time}秒ごとに記録)"

        ws.cell(row=1, column=2).value = f"終了時間({pitch_time}秒ごとに記録)"
        ws.cell(row=1, column=3).value = f"経過時間({pitch_time}秒ごとに記録)"
        ws.cell(row=1, column=4).value = "タイトル"
        ws.cell(row=1, column=5).value = "アプリケーション名"
        ws.cell(row=1, column=6).value = "アイドル時間(秒)"
        ws.cell(row=1, column=7).value = "ログインid"
        ws.cell(row=1, column=8).value = "コンピュータ名"

        wb.save(out_file)
        # # test1.txt を読み取り専用にする
        Path(out_file).chmod(0o444)

    try:
        while True:
            new_hwnd, new_time, new_title, new_exe, new_idl = get_active_window_info()
            wb = openpyxl.load_workbook(out_file)
            ws = wb[sheet_name]
            if last_hwnd != 0:  # 2回目以降
                # 最終行を更新
                max_row = ws.max_row
                dt_deff = new_time - last_time
                ws.cell(row=max_row, column=2).value = new_time  # 終了時間
                ws.cell(row=max_row, column=3).value = dt_deff  # 経過時間
                if new_idl > 0:
                    ws.cell(row=max_row, column=6).value = ws.cell(
                        row=max_row, column=6
                    ).value + (
                        new_idl - last_idl
                    )  # アイドル時間
            if new_hwnd != last_hwnd:
                # 新たな行を追加
                max_row = ws.max_row + 1
                ws.cell(row=max_row, column=1).value = new_time  # 開始時間
                ws.cell(row=max_row, column=2).value = new_time  # 終了時間
                ws.cell(row=max_row, column=3).value = 0  # 経過時間
                ws.cell(row=max_row, column=4).value = new_title  # ウィンドウタイトル
                ws.cell(row=max_row, column=5).value = new_exe  # exe名
                ws.cell(row=max_row, column=6).value = 0  # アイドル時間
                ws.cell(row=max_row, column=7).value = login_id  # ログオン名
                ws.cell(row=max_row, column=8).value = pc_NAME  # コンピュータ名
                # 今回の値を保存
                last_hwnd, last_time = (
                    new_hwnd,
                    new_time,
                )  # ウィンドウハンドル、開始時間を保存

            # # test2.txt の読み取り専用を外す
            Path(out_file).chmod(0o644)
            wb.save(out_file)
            # # test1.txt を読み取り専用にする
            Path(out_file).chmod(0o444)
            last_idl = new_idl  # アイドル時間を保存
            time.sleep(pitch_time)

    except KeyboardInterrupt:
        print("FINISH!")


layout = [
    [eg.Text("PC業務の記録を開始します。", key="-MAIN-")],
    [eg.Button("記録開始", key="-START-"), eg.Button("キャンセル", key="-CANCEL-")],
]

window = eg.Window(TITLE, layout, size=(400, 150))

while True:
    event, values = window.read()
    if event == eg.WIN_CLOSED or event == "-CANCEL-":
        eg.popup("PC業務の記録を中断しました。")
        break
    elif event == "-START-":
        threading.Thread(target=record_loop, daemon=True).start()
        window["-MAIN-"].update(
            "PC業務記録中です。。。\n\n記録データは、\nデスクトップの「PC業務記録データ」フォルダ\nの中に保存されています。"
        )
        window["-CANCEL-"].update("記録中断")
        window["-START-"].update(disabled=True)
